import os
import openpyxl
from platform import python_version
from datetime import datetime
starting_time = datetime.now()

ver = python_version()

os.system("cls")

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")

inputAttendance = "input_attendance.csv"
inputRegisteredFile = "input_registered_students.csv"

roll_to_name = {}
roll_attendance = {}
dates = []

def consolidate_attendance_func():
    try:
        outputFileName = "output/attendance_report_consolidated.xlsx"
        outputFile = openpyxl.Workbook()
        outputSheet = outputFile.active

        outputSheet.cell(row=1, column=1).value = "Roll"
        outputSheet.cell(row=1, column=2).value = "Name"

        last = -1
        for i, date in enumerate(dates):
            outputSheet.cell(row=1, column=3+i).value = date
            last = i+3

        last += 1
        list = ["Actual Lecture Taken", "Total Real", "% Attendance"]

        for i, title in enumerate(list):
            outputSheet.cell(row=1, column=last+i).value = title

        for i, rollNum in enumerate(roll_to_name.keys()):
            outputSheet.cell(row=i+2, column=1).value = rollNum
            outputSheet.cell(row=i+2, column=2).value = roll_to_name[rollNum]

            present = 0
            for j, date in enumerate(dates):
                if date not in roll_attendance[rollNum]:
                    outputSheet.cell(row=i+2, column=j+3).value = "A"
                else:
                    list = roll_attendance[rollNum][date]
                    total = list[0]+list[1]+list[2]
                    if total == 0:
                        outputSheet.cell(row=i+2, column=j+3).value = "A"
                    else:
                        outputSheet.cell(row=i+2, column=j+3).value = "P"
                        present += 1

            outputSheet.cell(row=i+2, column=last).value = len(dates)
            outputSheet.cell(row=i+2, column=last+1).value = present
            percentage_attendance = (100*present)/len(dates)
            percentage_attendance = round(percentage_attendance, 2)
            outputSheet.cell(row=i+2, column=last+2).value = percentage_attendance

        outputFile.save(outputFileName)
    except:
        print("Folder output does not exist")
        exit()

def roll_attendance_func():
    title = ["Date", "Roll", "Name", "Total Attendance Count",
             "Real", "Duplicate", "Invalid", "Absent"]

    for rollNum in roll_to_name.keys():
        try:
            outputFileName = "output/" + rollNum + ".xlsx"
            outputFile = openpyxl.Workbook()
            outputSheet = outputFile.active

            for i, word in enumerate(title):
                outputSheet.cell(row=1, column=i+1).value = word
            outputSheet.cell(row=2, column=2).value = rollNum
            outputSheet.cell(row=2, column=3).value = roll_to_name[rollNum]

            attendance = roll_attendance[rollNum]  # map of date -> array

            for i, date in enumerate(attendance.keys()):
                outputSheet.cell(row=3+i, column=1).value = date
                list = attendance[date]
                total = list[0]+list[1]+list[2]

                outputSheet.cell(row=3+i, column=4).value = total
                outputSheet.cell(row=3+i, column=5).value = list[0]
                outputSheet.cell(row=3+i, column=6).value = list[1]
                outputSheet.cell(row=3+i, column=7).value = list[2]

                if total == 0:
                    outputSheet.cell(row=3+i, column=8).value = 1
                else:
                    outputSheet.cell(row=3+i, column=8).value = 0

            outputFile.save(outputFileName)
        except:
            print("The output folder does not exist")
            exit()

